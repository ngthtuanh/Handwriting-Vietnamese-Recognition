import streamlit as st
import torch
import pandas as pd
from vietocr.tool.predictor import Predictor
from vietocr.tool.config import Cfg
from PIL import Image
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
import os

# Load OCR model
config = Cfg.load_config_from_file('./config.yml')
config['weights'] = './transformerocr.pth'
config['device'] = 'cuda' if torch.cuda.is_available() else 'cpu'
detector = Predictor(config)

# Streamlit UI Configuration
st.set_page_config(page_title="Handwritten Vietnamese", page_icon="📝", layout="wide")

# Initialize session state variables
if "page" not in st.session_state:
    st.session_state.page = "INTRODUCE"
if "results" not in st.session_state:
    st.session_state.results = []

# Main title
st.markdown("<h1 style='text-align: center;'>📝 HANDWRITTEN VIETNAMESE CHARACTER RECOGNITION - LINE LEVEL</h1>", unsafe_allow_html=True)

def switch_page(target_page):
    st.session_state.page = target_page
    
# Function to handle file uploads and OCR processing
def process_files(uploaded_files):
    st.session_state.results = []
    cols = st.columns(min(3, len(uploaded_files)))
    
    for idx, uploaded_file in enumerate(uploaded_files):
        image = Image.open(uploaded_file).convert('RGB')
        with cols[idx % len(cols)]:  
            st.image(image, caption=f"📷 {uploaded_file.name}", use_column_width=True)
        
        with st.spinner(f"🔍 Recognizing {uploaded_file.name}..."):
            prediction = detector.predict(image)
        
        img_byte_arr = BytesIO()
        image.save(img_byte_arr, format='PNG')
        st.session_state.results.append({"Filename": uploaded_file.name, "Prediction": prediction, "Image": img_byte_arr.getvalue()})
    
    st.write("### 📄 Recognition Results")
    df = pd.DataFrame(st.session_state.results, columns=["Filename", "Prediction"])
    st.table(df)

# Function to export results to Excel
def export_to_excel():
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["Filename", "Prediction", "Image"])
    
    for idx, result in enumerate(st.session_state.results, start=2):
        ws[f"A{idx}"] = result["Filename"]
        ws[f"B{idx}"] = result["Prediction"]
        img = openpyxl.drawing.image.Image(BytesIO(result["Image"]))
        img.width, img.height = 100, 50
        ws.add_image(img, f"C{idx}")
    
    wb.save(output)
    output.seek(0)
    
    st.download_button(
        label="📥 Download Excel file",
        data=output,
        file_name="Results_with_Images.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# # Navigation buttons
col_empty, col1, col2, col_empty2 = st.columns([2, 4, 2, 2])
with col1:
    if st.button("🏠 INTRODUCE"):
        switch_page("INTRODUCE")
with col2:
    if st.button("✍️ VIETNAMESE HANDWRITTING RESULT"):
        switch_page("VIETNAMESE HANDWRITTING RESULT")

if st.session_state.page == "INTRODUCE":
    with st.container():
        col_text, col_image = st.columns([1, 1])
        with col_text:
            st.write("## About Handwritten Vietnamese")
            st.write("The Handwritten Vietnamese application helps recognize Vietnamese handwriting at the line level, enabling fast and accurate document digitization.")
            
            st.write("### Key Features:")
            st.write("- Automatic handwriting recognition with good accuracy.")
            st.write("- Support many input image formats.")
            st.write("- Export recognition results to Excel file along with the original image.")

            st.write("### How to Use:")
            st.write("1. Select '✍️ VIETNAMESE HANDWRITTING RESULT' to upload handwritten image.")
            st.write("2. The system will automatically recognize and display the result.")
            st.write("3. Download the result as an Excel file for storage.")

            st.write("### Model Evaluation:")
            st.write("- **Character Accuracy (1 - CER):** 90.10%")
            st.write("- **Word Accuracy (1 - WER):** 76.07%")
            st.write("- **CER (Character Error Rate):** 0.0990")
            st.write("- **WER (Word Error Rate):** 0.2393")
            
            st.write("### Overall Assessment:")
            st.write("- The model performs well with high character accuracy, making it suitable for Vietnamese handwriting recognition at the line level.")
            st.write("- However, the word error rate remains relatively high, which may affect full-word recognition accuracy.")
        
        with col_image:
            sample_images = ["./data_line/InkData_line_processed/20160411_0082_9129_1_tg_1_3.png",
                              "./data_line/InkData_line_processed/20151214_0065_26558_2_tg_1_7.png", 
                              "./data_line/InkData_line_processed/20160113_0069_28472_3_tg_2_6.png", 
                              "./data_line/InkData_line_processed/20160604_0199_25463_tg_0_3.png"]
            for img_path in sample_images:
                image = Image.open(img_path).convert('RGB')
                st.image(image, caption=f"Example: {img_path}", use_column_width=True)
                prediction = detector.predict(image)
                st.write(f"**📄 Recognition Results: ** {prediction}")

elif st.session_state.page == "VIETNAMESE HANDWRITTING RESULT":
    new_files = st.file_uploader("📂 Upload new handwritten images", type=["png", "jpg", "jpeg"], accept_multiple_files=True)
    if new_files:
        process_files(new_files)

    # Nút xuất file Excel hiển thị trên cả hai trang nếu có kết quả
    if st.session_state.results:
        export_to_excel()
